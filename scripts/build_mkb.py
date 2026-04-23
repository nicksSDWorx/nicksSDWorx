"""
Builder for Payroll Discovery Document MKB (v2, Dutch-only).
Phased build.
Phase 5 status: all 7 content sheets filled.
Phase 8 status: SD Worx branding applied (logo, brand palette, refined layout).
"""
import os
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties

OUTPUT = "/home/user/nicksSDWorx/outputs/Payroll_Discovery_Document_MKB.xlsx"
LOGO_PATH = "/home/user/nicksSDWorx/assets/sdworx_logo_fc_crop.png"

# --- SD Worx brand palette ---
BRAND_NAVY = "203040"
BRAND_RED = "E01030"
BRAND_YELLOW = "F0B000"
BRAND_BLUE = "5080A0"
BRAND_NAVY_SOFT = "3B4B64"
PAGE_BG = "FFFFFF"
SUB_FILL_HEX = "E8ECF2"
INPUT_FILL_HEX = "FFF4D1"
DEFAULT_FILL_HEX = "F4F5F7"
INFO_FILL_HEX = "FFF8E2"
BORDER_GREY = "C5CCD6"
HELP_GREY = "6B7684"
DEFAULT_TEXT = "5F6B7A"

# --- Fills ---
HEADER_FILL = PatternFill("solid", fgColor=PAGE_BG)
ACCENT_STRIPE_FILL = PatternFill("solid", fgColor=BRAND_YELLOW)
SUBHEADER_FILL = PatternFill("solid", fgColor=SUB_FILL_HEX)
INPUT_FILL = PatternFill("solid", fgColor=INPUT_FILL_HEX)
DEFAULT_FILL = PatternFill("solid", fgColor=DEFAULT_FILL_HEX)
INFO_FILL = PatternFill("solid", fgColor=INFO_FILL_HEX)

# --- Fonts ---
HEADER_FONT = Font(name="Calibri", size=18, bold=True, color=BRAND_NAVY)
SUBHEADER_FONT = Font(name="Calibri", size=12, bold=True, color=BRAND_NAVY)
BASE_FONT = Font(name="Calibri", size=11, color=BRAND_NAVY)
LABEL_FONT = Font(name="Calibri", size=11, color=BRAND_NAVY)
REQ_FONT = Font(name="Calibri", size=11, color=BRAND_NAVY, bold=False)
REQ_STAR_FONT = Font(name="Calibri", size=11, color=BRAND_RED, bold=True)
HELP_FONT = Font(name="Calibri", size=9, italic=True, color=HELP_GREY)
DEFAULT_FONT = Font(name="Calibri", size=11, italic=True, color=DEFAULT_TEXT)
INTRO_FONT = Font(name="Calibri", size=11, color=BRAND_NAVY)
INTRO_BOLD_FONT = Font(name="Calibri", size=12, bold=True, color=BRAND_NAVY)

THIN = Side(border_style="thin", color=BORDER_GREY)
BORDER_INPUT = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

UNLOCKED = Protection(locked=False)
LOCKED = Protection(locked=True)

# --- Sheet structure ---
SHEETS = [
    "Start",
    "1. Bedrijf & Contact",
    "2. Payroll basis",
    "3. Loonheffing",
    "4. Reserveringen & Pensioen",
    "5. Verlof & Grootboek",
    "6. Akkoord",
]
SHEET_TITLES = {
    "Start": "Payroll Discovery Document - MKB",
    "1. Bedrijf & Contact": "1. Bedrijf & Contact",
    "2. Payroll basis": "2. Payroll basis",
    "3. Loonheffing": "3. Loonheffing",
    "4. Reserveringen & Pensioen": "4. Reserveringen & Pensioen",
    "5. Verlof & Grootboek": "5. Verlof & Grootboek",
    "6. Akkoord": "6. Akkoord",
}

VALIDATIONS = {
    "val_JaNee": ["Ja", "Nee"],
    "val_Salarisperiode": ["Week", "4 weken", "Maand"],
    "val_Taal": ["Nederlands", "Engels"],
    "val_Indeling": ["Klein", "Middel", "Groot"],
    "val_Oproep": [
        "Nee", "Ja, met verplichting te komen",
        "Ja, zonder verplichting te komen", "Ja, beide",
    ],
    "val_Maand": [
        "Januari", "Februari", "Maart", "April", "Mei", "Juni",
        "Juli", "Augustus", "September", "Oktober", "November", "December",
        "Periode 1", "Periode 2", "Periode 3", "Periode 4", "Periode 5",
        "Periode 6", "Periode 7", "Periode 8", "Periode 9", "Periode 10",
        "Periode 11", "Periode 12", "Periode 13",
    ],
    "val_Zichtbaarheid": ["Op de betaaldatum", "# dagen na de betaaldatum"],
    "val_Formaat": ["XML", "Excel", "CSV", "TXT", "Anders"],
    "val_Basisperiode": [
        "In de periode", "In de maand", "Per 1 januari", "Per 31 december",
        "Geboren voor", "Geboren op of na", "AOW Gerechtigd",
    ],
    "val_FinSysteem": [
        "Exact Online", "AFAS", "Twinfield", "Visma.net", "SnelStart", "Unit4", "Anders",
    ],
    "val_30pct": ["Nee", "Ja, bruto-aftopping", "Ja, netto-vergoeding"],
    "val_Toggle": ["Basis", "Uitgebreid"],
    "val_Checkbox": ["☐", "☑"],
    "val_Sector": [
        "1 - Agrarisch bedrijf", "2 - Tabakverwerkende industrie",
        "3 - Bouwbedrijf", "4 - Baggerbedrijf",
        "5 - Hout en emballage-industrie", "6 - Timmerindustrie",
        "7 - Meubel- en orgelbouw industrie",
        "8 - Groothandel hout, zagerijen, schaverijen en houtbereidingsindustrie",
        "9 - Grafische industrie", "10 - Metaal industrie",
        "11 - Elektronische industrie", "12 - Metaal- en technische bedrijfstakken",
        "13 - Bakkerijen", "14 - Suikerverwerkende industrie",
        "15 - Slagersbedrijven", "16 - Slagers overig",
        "17 - Detailhandel en ambachten", "18 - Reiniging",
        "19 - Grootwinkelbedrijf", "20 - Havenbedrijven",
        "21 - Havenclassificeerders", "22 - Binnenscheepvaart",
        "23 - Visserij", "24 - Koopvaardij",
        "25 - Vervoer KLM", "26 - Vervoer NS",
        "27 - Vervoer posterijen", "28 - Taxivervoer",
        "29 - Openbaar vervoer", "30 - Besloten busvervoer",
        "31 - Overig personenvervoer te land en in de lucht",
        "32 - Overig goederenvervoer te land en in de lucht",
        "33 - Horeca algemeen", "34 - Horeca catering",
        "35 - Gezondheid, geestelijke en maatschappelijke belangen",
        "38 - Banken", "39 - Verzekeringswezen en ziekenfondsen",
        "40 - Uitgeverij", "41 - Groothandel I", "42 - Groothandel II",
        "43 - Zakelijke dienstverlening I", "44 - Zakelijke dienstverlening II",
        "45 - Zakelijke dienstverlening III", "46 - Zuivelindustrie",
        "47 - Textielindustrie",
        "48 - Steen-, cement-, glas- en keramische industrie",
        "49 - Chemische industrie", "50 - Voedingsindustrie",
        "51 - Algemene industrie", "52 - Uitzendbedrijven",
        "53 - Bewakingsondernemingen", "54 - Culturele instellingen",
        "55 - Overige takken van bedrijf en beroep",
        "56 - Schildersbedrijf", "57 - Stukadoorsbedrijf",
        "58 - Dakdekkersbedrijf", "59 - Mortelbedrijf",
        "60 - Steenhouwersbedrijf",
        "61 - Overheid, onderwijs en wetenschappen",
        "62 - Overheid, rijk, politie en rechterlijke macht",
        "63 - Overheid, defensie",
        "64 - Overheid, provincies, gemeenten en waterschappen",
        "65 - Overheid, openbare nutsbedrijven",
        "66 - Overheid, overige instellingen",
        "67 - Werk en (re)integratie",
        "68 - Railbouw", "69 - Telecommunicatie",
    ],
}

# --- Global registries ---
REQUIRED_REFS = []  # list of (sheet_name, coord)
DV_CACHE = {}       # (sheet_name, list_name) -> DataValidation


# --- Helpers ---
def setup_sheet(ws, title):
    ws.column_dimensions["A"].width = 42
    for col in ("B", "C", "D", "E", "F", "G", "H"):
        ws.column_dimensions[col].width = 28
    # Branded header: white background, logo top-left, title right
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = title
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal="right", vertical="center", indent=2)
    ws.row_dimensions[1].height = 55
    # Accent stripe row 2 (brand yellow) as a thin divider
    for col_idx in range(1, 9):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = ACCENT_STRIPE_FILL
    ws.row_dimensions[2].height = 4
    # Embed logo anchored to A1 (floats left within header row)
    img = Image(LOGO_PATH)
    img.width = 190
    img.height = 63
    img.anchor = "A1"
    ws.add_image(img)
    # View + print
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.protection.sheet = True


def subheader(ws, row, text, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = SUBHEADER_FILL
    c.font = SUBHEADER_FONT
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = Border(left=Side(border_style="thick", color=BRAND_YELLOW))
    ws.row_dimensions[row].height = 24


def label(ws, row, text, required=False, col=1):
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    c = ws.cell(row=row, column=col)
    if required:
        navy = InlineFont(rFont="Calibri", sz=11, color=BRAND_NAVY)
        red = InlineFont(rFont="Calibri", sz=11, b=True, color=BRAND_RED)
        c.value = CellRichText([TextBlock(navy, text), TextBlock(red, "  *")])
    else:
        c.value = text
    c.font = LABEL_FONT
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 18, 18)


def plain(ws, coord, text, font=None, alignment=None):
    c = ws[coord]
    c.value = text
    c.font = font or BASE_FONT
    if alignment:
        c.alignment = alignment


def input_cell(ws, coord, required=False, comment=None, default=None, is_default=False):
    c = ws[coord]
    if default is not None:
        c.value = default
    c.fill = DEFAULT_FILL if is_default else INPUT_FILL
    c.font = DEFAULT_FONT if is_default else BASE_FONT
    c.border = BORDER_INPUT
    c.protection = UNLOCKED
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    if comment:
        c.comment = Comment(comment, "SD Worx")
    if required:
        REQUIRED_REFS.append((ws.title, coord))


def dropdown(ws, coord, listname, required=False, comment=None, default=None, is_default=False):
    input_cell(ws, coord, required=required, comment=comment, default=default, is_default=is_default)
    key = (ws.title, listname)
    dv = DV_CACHE.get(key)
    if dv is None:
        dv = DataValidation(
            type="list", formula1=f"={listname}", allow_blank=True, showDropDown=False
        )
        dv.error = "Ongeldige keuze. Kies uit de lijst."
        dv.errorTitle = "Ongeldige invoer"
        ws.add_data_validation(dv)
        DV_CACHE[key] = dv
    dv.add(coord)


def help_line(ws, coord, text):
    c = ws[coord]
    c.value = text
    c.font = HELP_FONT
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def opmerkingen_block(ws, start_row, rows=4, span=8):
    subheader(ws, start_row, "Opmerkingen", span=span)
    ws.merge_cells(
        start_row=start_row + 1, start_column=1,
        end_row=start_row + rows, end_column=span,
    )
    c = ws.cell(row=start_row + 1, column=1)
    c.fill = INPUT_FILL
    c.border = BORDER_INPUT
    c.protection = UNLOCKED
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)


def footnote(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = HELP_FONT


# --- Validations sheet ---
def build_validations(wb):
    vs = wb.create_sheet("_Validations")
    for col_idx, (name, values) in enumerate(VALIDATIONS.items(), start=1):
        letter = get_column_letter(col_idx)
        vs.column_dimensions[letter].width = 40
        vs.cell(row=1, column=col_idx, value=name).font = Font(bold=True)
        for i, v in enumerate(values, start=2):
            vs.cell(row=i, column=col_idx, value=v)
        ref = f"'_Validations'!${letter}$2:${letter}${len(values) + 1}"
        wb.defined_names[name] = DefinedName(name=name, attr_text=ref)
    vs.sheet_state = "hidden"


# --- Start sheet ---
def build_start(wb):
    ws = wb["Start"]
    # Intro
    ws.merge_cells("A3:H3")
    plain(ws, "A3", "Welkom! Dit document is de basis voor de inrichting van uw loonadministratie in Cobra (SD Worx).",
          font=Font(name="Calibri", size=12, bold=True))
    ws.merge_cells("A4:H4")
    plain(ws, "A4",
          "Vul per tabblad de gegevens in. Geschatte invultijd: 30 - 45 minuten. Velden met * zijn verplicht.",
          font=INTRO_FONT)
    ws.merge_cells("A5:H5")
    plain(ws, "A5",
          "Geel = door u in te vullen.   Lichtgrijs = voorstel (overschrijfbaar).   Bij twijfel: leeglaten en contact opnemen met uw consultant.",
          font=HELP_FONT)

    # Consultant block
    subheader(ws, 7, "Uw contactpersoon bij SD Worx")
    label(ws, 8, "Naam consultant")
    input_cell(ws, "B8", comment="In te vullen door SD Worx-consultant.")
    label(ws, 9, "E-mail")
    input_cell(ws, "B9", comment="In te vullen door SD Worx-consultant.")
    label(ws, 10, "Telefoon")
    input_cell(ws, "B10", comment="In te vullen door SD Worx-consultant.")

    # Bijlagen checklist
    subheader(ws, 12, "Benodigde bijlagen  (aanvinken wat is meegestuurd)")
    attachments = [
        "Kopie laatste loonstrook van ca. 5 medewerkers (incl. 1 met variabele beloning, 1 met pensioen)",
        "Kopie WHK-beschikking lopend jaar",
        "Kopie grootboekschema + laatste journaalpost (alleen als u GL-export wilt)",
        "Kopie pensioenregeling (alleen als u NIET bij een bedrijfstakpensioenfonds zit)",
        "CAO (alleen als van toepassing)",
        "Functielijst (functiecode + omschrijving)",
    ]
    for i, text in enumerate(attachments):
        r = 13 + i
        dropdown(ws, f"A{r}", "val_Checkbox", default="☐", is_default=True)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        tc = ws.cell(row=r, column=2, value=text)
        tc.font = BASE_FONT
        tc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

    # Progress indicator
    subheader(ws, 20, "Voortgang verplichte velden")
    label(ws, 21, "Aantal ingevuld / totaal")
    # Placeholder; filled at end
    c = ws["B21"]
    c.fill = DEFAULT_FILL
    c.font = DEFAULT_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws["B21"].value = "(wordt berekend)"
    label(ws, 22, "Percentage ingevuld")
    c = ws["B22"]
    c.fill = DEFAULT_FILL
    c.font = DEFAULT_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws["B22"].value = "(wordt berekend)"
    help_line(ws, "C21",
              "Tip: sla het bestand op of druk op F9 om de teller te verversen.")

    opmerkingen_block(ws, 24)


# --- 1. Bedrijf & Contact ---
def build_bedrijf(wb):
    ws = wb["1. Bedrijf & Contact"]

    subheader(ws, 3, "Bedrijfsgegevens")
    rows_company = [
        ("Bedrijfsnaam", "B4", True, None),
        ("Straat + huisnummer", "B5", True, None),
        ("Postcode", "B6", True, "Formaat: 1234 AB"),
        ("Plaats", "B7", True, None),
        ("KvK-nummer", "B8", False, "8 cijfers, zonder vestigingsnummer."),
        ("Loonheffingsnummer", "B9", False,
         "Formaat: 123456789L01. Wordt ook gevraagd op tab 3. Loonheffing; hier mag u overslaan."),
    ]
    for i, (lab, cell, req, cmt) in enumerate(rows_company):
        label(ws, 4 + i, lab, required=req)
        input_cell(ws, cell, required=req, comment=cmt)

    subheader(ws, 11, "Contactpersoon HR / salarisadministratie")
    rows_hr = [
        ("Naam", "B12", True),
        ("Functie", "B13", False),
        ("Telefoon", "B14", True),
        ("E-mail", "B15", True),
    ]
    for i, (lab, cell, req) in enumerate(rows_hr):
        label(ws, 12 + i, lab, required=req)
        input_cell(ws, cell, required=req)

    subheader(ws, 17, "Contactpersoon Finance / boekhouding")
    rows_fin = [
        ("Naam", "B18", True),
        ("Functie", "B19", False),
        ("Telefoon", "B20", True),
        ("E-mail", "B21", True),
    ]
    for i, (lab, cell, req) in enumerate(rows_fin):
        label(ws, 18 + i, lab, required=req)
        input_cell(ws, cell, required=req)

    subheader(ws, 23, "Bankgegevens (voor SEPA-betaalbestand nettolonen)")
    label(ws, 24, "IBAN-rekeningnummer", required=True)
    input_cell(ws, "B24", required=True,
               comment="Nederlands IBAN, bijv. NL91 ABNA 0417 1643 00. Spaties mogen, worden genegeerd.")
    label(ws, 25, "Naam bank")
    input_cell(ws, "B25")
    label(ws, 26, "BIC / SWIFT-code")
    input_cell(ws, "B26",
               comment="Alleen invullen als de bank NIET in de IBAN-zone (Europa) zit.")

    footnote(ws, 28, "* = verplicht veld. Niet-verplichte velden mogen leeg blijven.")
    opmerkingen_block(ws, 30)


# --- 2. Payroll basis ---
def build_payroll(wb):
    ws = wb["2. Payroll basis"]

    subheader(ws, 3, "CAO (Collectieve arbeidsovereenkomst)")
    label(ws, 4, "Is een CAO van toepassing?", required=True)
    dropdown(ws, "B4", "val_JaNee", required=True)
    label(ws, 5, "Zo ja, welke CAO?")
    input_cell(ws, "B5", comment="Voluit de officiele naam, bv. 'CAO Metaal en Techniek'.")

    subheader(ws, 7, "Werktijden (standaard fulltime)")
    label(ws, 8, "Uren per week", required=True)
    input_cell(ws, "B8", required=True, default=40, is_default=True,
               comment="Standaard fulltime = 40 uur. Wijk alleen af als uw CAO iets anders voorschrijft.")
    label(ws, 9, "Uren per dag (maandag - vrijdag)")
    days = [("C9", "ma"), ("D9", "di"), ("E9", "wo"), ("F9", "do"), ("G9", "vr")]
    for coord, day in days:
        # day label above the input
        top = ws.cell(row=8, column=ws[coord].column, value=day)
        top.font = HELP_FONT
        top.alignment = Alignment(horizontal="center", vertical="center")
        input_cell(ws, coord, default=8, is_default=True)
        ws[coord].alignment = Alignment(horizontal="center", vertical="center")

    subheader(ws, 11, "Salarisperiode")
    label(ws, 12, "Salarisperiode", required=True)
    dropdown(ws, "B12", "val_Salarisperiode", required=True, default="Maand", is_default=True,
             comment="Voor MKB bijna altijd 'Maand'. '4 weken' enkel bij CAO-verplichting.")

    subheader(ws, 14, "Salarisstrook")
    label(ws, 15, "Taal salarisstrook", required=True)
    dropdown(ws, "B15", "val_Taal", required=True, default="Nederlands", is_default=True)
    label(ws, 16, "Logo van uw bedrijf op salarisstrook?", required=True)
    dropdown(ws, "B16", "val_JaNee", required=True,
             comment="Bij 'Ja': stuur het logo als PNG of JPG mee (min. 300 dpi).")
    label(ws, 17, "Strook zichtbaar voor medewerker", required=True)
    dropdown(ws, "B17", "val_Zichtbaarheid", required=True,
             default="Op de betaaldatum", is_default=True)
    label(ws, 18, "Aantal dagen na betaaldatum")
    input_cell(ws, "B18",
               comment="Alleen invullen als hierboven '# dagen na de betaaldatum' is gekozen.")

    subheader(ws, 20, "30%-regeling (voor buitenlandse kenniswerkers)")
    label(ws, 21, "30%-regeling van toepassing?", required=True)
    dropdown(ws, "B21", "val_30pct", required=True, default="Nee", is_default=True,
             comment="Kies 'Nee' als u geen buitenlandse kenniswerkers in dienst heeft.")
    label(ws, 22, "Aantal medewerkers met 30%-regeling")
    input_cell(ws, "B22", comment="Alleen invullen bij 'Ja' hierboven.")

    subheader(ws, 24, "Overige werknemerstypen")
    label(ws, 25, "Stagiaires in dienst?", required=True)
    dropdown(ws, "B25", "val_JaNee", required=True, default="Nee", is_default=True)
    label(ws, 26, "Oproepkrachten in dienst?", required=True)
    dropdown(ws, "B26", "val_Oproep", required=True, default="Nee", is_default=True,
             comment="Kies het type oproepovereenkomst. Bij meerdere soorten: kies 'Ja, beide'.")

    opmerkingen_block(ws, 28)


# --- 3. Loonheffing ---
def build_loonheffing(wb):
    ws = wb["3. Loonheffing"]

    subheader(ws, 3, "Gegevens aangifte loonheffingen")
    label(ws, 4, "Loonheffingsnummer", required=True)
    input_cell(ws, "B4", required=True,
               comment="Formaat: 123456789L01. Staat op uw aangifte loonheffingen of in uw portaal bij de Belastingdienst.")
    label(ws, 5, "Sector (sectorindeling Belastingdienst)", required=True)
    dropdown(ws, "B5", "val_Sector", required=True,
             comment="Kies de sector waarin uw bedrijf is ingedeeld. Bij twijfel: zie de sectorbeschikking van de Belastingdienst.")
    label(ws, 6, "Indeling werkgever", required=True)
    dropdown(ws, "B6", "val_Indeling", required=True,
             comment="Klein: minder dan 25 werknemers. Middel: 25 - 100 werknemers. Groot: meer dan 100 werknemers. Bij twijfel: leeglaten, wij bepalen dit.")
    label(ws, 7, "CBS CAO-code")
    input_cell(ws, "B7",
               comment="Optioneel. De 4-cijferige code van uw CAO volgens het CBS. Bij geen CAO: leeglaten.")
    label(ws, 8, "WBSO-subsidie van toepassing?", required=True)
    dropdown(ws, "B8", "val_JaNee", required=True, default="Nee", is_default=True,
             comment="WBSO = Wet Bevordering Speur- en Ontwikkelingswerk. Alleen 'Ja' als u een WBSO-beschikking van RVO heeft.")

    subheader(ws, 10, "WHK-premie (gedifferentieerde premie Werkhervattingskas)")
    ws.merge_cells("A11:H14")
    c = ws["A11"]
    c.value = (
        "U hoeft hier GEEN premiepercentages in te vullen.\n\n"
        "Stuur de WHK-beschikking van de Belastingdienst voor het lopende jaar mee "
        "als bijlage (zie de bijlagen-checklist op het tabblad 'Start').\n\n"
        "Wij verwerken op basis van die beschikking de WGA-premie, ZW-premie "
        "en eventueel eigenrisicodragerschap in de inrichting."
    )
    c.font = BASE_FONT
    c.fill = INFO_FILL
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.border = BORDER_INPUT

    opmerkingen_block(ws, 16)


# --- 4. Reserveringen & Pensioen ---
def build_reserveringen(wb):
    ws = wb["4. Reserveringen & Pensioen"]

    subheader(ws, 3, "Vakantiegeld")
    label(ws, 4, "Uitbetalingspercentage")
    input_cell(ws, "B4", default="8,33%", is_default=True,
               comment="Wettelijk minimum 8,00%. In veel CAO's 8,33% (= 1 maand extra salaris).")
    label(ws, 5, "Referentieperiode van")
    input_cell(ws, "B5", default="Juni", is_default=True)
    label(ws, 6, "Referentieperiode tot en met")
    input_cell(ws, "B6", default="Mei", is_default=True)
    label(ws, 7, "Uitbetaalmaand")
    dropdown(ws, "B7", "val_Maand", default="Mei", is_default=True)

    subheader(ws, 9, "13e maand")
    label(ws, 10, "13e maand van toepassing?", required=True)
    dropdown(ws, "B10", "val_JaNee", required=True, default="Nee", is_default=True)
    label(ws, 11, "Uitbetaalmaand 13e maand")
    dropdown(ws, "B11", "val_Maand",
             comment="Alleen invullen als '13e maand van toepassing' = Ja. Meestal december.")

    subheader(ws, 13, "Pensioen")
    label(ws, 14, "Aangesloten bij bedrijfstakpensioenfonds (BPF)?", required=True)
    dropdown(ws, "B14", "val_JaNee", required=True,
             comment="BPF = verplicht pensioenfonds per branche (bv. PFZW, BPF Bouw, BPF Metaal). Bij 'Ja' volgt alles uit het BPF-reglement. Bij 'Nee' heeft u een eigen regeling.")

    # Branch A: BPF
    plain(ws, "A16", "Bij BPF = Ja", font=Font(name="Calibri", size=10, bold=True, color="1F3864"))
    label(ws, 17, "Naam bedrijfstakpensioenfonds")
    input_cell(ws, "B17", comment="Bijv. 'PFZW', 'BPF Metaal en Techniek', 'BPF Bouw'.")
    label(ws, 18, "Aansluitingsnummer bij het fonds")
    input_cell(ws, "B18")

    # Branch B: eigen regeling
    plain(ws, "A20", "Bij BPF = Nee (eigen pensioenregeling)",
          font=Font(name="Calibri", size=10, bold=True, color="1F3864"))
    label(ws, 21, "Premie werknemer (%)")
    input_cell(ws, "B21")
    label(ws, 22, "Premie werkgever (%)")
    input_cell(ws, "B22")
    label(ws, 23, "Franchise (EUR per jaar)")
    input_cell(ws, "B23",
               comment="Bedrag waarover GEEN pensioen wordt opgebouwd. Bij BPF: leeglaten. Bij twijfel: vraag uw pensioenadviseur.")
    label(ws, 24, "Maximum pensioengevend jaarloon (EUR)")
    input_cell(ws, "B24",
               comment="Wettelijk maximum 2025: circa EUR 137.800. Laat leeg voor 'geen maximum'.")
    label(ws, 25, "Minimum leeftijd voor pensioenopbouw")
    input_cell(ws, "B25", default=21, is_default=True)
    label(ws, 26, "Maximum leeftijd voor pensioenopbouw")
    input_cell(ws, "B26", default="AOW-leeftijd", is_default=True,
               comment="Standaard tot AOW-leeftijd. Wijzig alleen als uw regeling anders bepaalt.")

    subheader(ws, 28, "Aanvullende regelingen")
    label(ws, 29, "Heeft u aanvullende regelingen?", required=True)
    dropdown(ws, "B29", "val_JaNee", required=True, default="Nee", is_default=True,
             comment="Denk aan WGA-hiaatverzekering, pensioen-excedent, netto pensioen, WIA-bodem. Bij 'Ja': uw consultant neemt dit mondeling met u door. Wij vragen hier geen details.")

    subheader(ws, 31, "Geavanceerd (optioneel) - Individueel Keuzebudget (IKB)")
    label(ws, 32, "IKB-toggle")
    dropdown(ws, "B32", "val_Toggle", default="Basis", is_default=True,
             comment="Kies 'Uitgebreid' als u een IKB-regeling heeft. Kies 'Basis' om IKB-velden te negeren.")
    label(ws, 33, "IKB-percentage")
    input_cell(ws, "B33", is_default=True,
               comment="Alleen invullen bij toggle = 'Uitgebreid'.")
    label(ws, 34, "IKB-referentieperiode van")
    input_cell(ws, "B34", is_default=True,
               comment="Alleen invullen bij toggle = 'Uitgebreid'.")
    label(ws, 35, "IKB-referentieperiode tot en met")
    input_cell(ws, "B35", is_default=True,
               comment="Alleen invullen bij toggle = 'Uitgebreid'.")
    label(ws, 36, "IKB-uitbetaalmaand")
    dropdown(ws, "B36", "val_Maand", is_default=True,
             comment="Alleen invullen bij toggle = 'Uitgebreid'.")

    opmerkingen_block(ws, 38)


# --- 5. Verlof & Grootboek ---
def build_verlof_gl(wb):
    ws = wb["5. Verlof & Grootboek"]

    subheader(ws, 3, "Wettelijke verlofregelingen (WAZO)")

    # Table header
    hdr_cells = [("A4", "Soort verlof"), ("B4", "Doorbetaald %"),
                 ("C4", "Pensioenopbouw"), ("D4", "Vakantiegeld-opbouw")]
    for coord, text in hdr_cells:
        c = ws[coord]
        c.value = text
        c.font = Font(name="Calibri", size=11, bold=True, color="1F3864")
        c.fill = SUBHEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 22

    wazo_rows = [
        ("Geboorteverlof (1 week)", 100, "Ja", "Ja"),
        ("Aanvullend geboorteverlof (5 weken)", 70, "Ja", "Nee"),
        ("Betaald ouderschapsverlof (9 weken)", 70, "Ja", "Nee"),
        ("Onbetaald ouderschapsverlof", 0, "Nee", "Nee"),
        ("Kortdurend zorgverlof", 70, "Ja", "Ja"),
        ("Langdurig zorgverlof", 0, "Nee", "Nee"),
        ("Onbetaald verlof (algemeen)", 0, "Nee", "Nee"),
    ]
    for i, (naam, pct, pens, vak) in enumerate(wazo_rows):
        r = 5 + i
        c = ws.cell(row=r, column=1, value=naam)
        c.font = BASE_FONT
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        input_cell(ws, f"B{r}", default=pct, is_default=True)
        ws[f"B{r}"].alignment = Alignment(horizontal="center", vertical="center")
        dropdown(ws, f"C{r}", "val_JaNee", default=pens, is_default=True)
        ws[f"C{r}"].alignment = Alignment(horizontal="center", vertical="center")
        dropdown(ws, f"D{r}", "val_JaNee", default=vak, is_default=True)
        ws[f"D{r}"].alignment = Alignment(horizontal="center", vertical="center")

    footnote(ws, 13,
             "Defaults = wettelijk minimum. Overschrijf indien uw CAO gunstigere voorwaarden biedt.")

    subheader(ws, 15, "Grootboek-export (journaalpost)")
    label(ws, 16, "Financieel systeem", required=True)
    dropdown(ws, "B16", "val_FinSysteem", required=True,
             comment="Kies uw financiele pakket. Kies 'Anders' als uw pakket er niet bij staat - wij nemen dan contact op.")
    label(ws, 17, "Bestandsformaat export", required=True)
    dropdown(ws, "B17", "val_Formaat", required=True,
             comment="Meestgebruikt: CSV of XML. Exact Online: XML. AFAS: XML of CSV.")
    label(ws, 18, "Journaalpost per kostenplaats splitsen?", required=True)
    dropdown(ws, "B18", "val_JaNee", required=True, default="Nee", is_default=True,
             comment="Kies 'Ja' als u per afdeling / kostenplaats gesplitste boekingen wilt.")

    ws.merge_cells("A20:H22")
    c = ws["A20"]
    c.value = ("Stuur uw GROOTBOEKSCHEMA + een recente JOURNAALPOST als bijlage mee "
               "(zie bijlagen-checklist op Start). Wij mappen uw looncomponenten hieraan - "
               "u hoeft hier geen aparte kolommen op te geven.")
    c.font = BASE_FONT
    c.fill = INFO_FILL
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.border = BORDER_INPUT

    opmerkingen_block(ws, 24)


# --- 6. Akkoord ---
def build_akkoord(wb):
    ws = wb["6. Akkoord"]

    ws.merge_cells("A3:H5")
    c = ws["A3"]
    c.value = ("Hierbij verklaart ondergetekende dat de inrichting van Cobra conform de gegevens "
               "in dit inventarisatiedocument is uitgevoerd en geeft akkoord op de onderstaande punten.")
    c.font = BASE_FONT
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)

    subheader(ws, 7, "Ondertekening")
    label(ws, 8, "Naam klant / bedrijf", required=True)
    input_cell(ws, "B8", required=True)
    label(ws, 9, "Naam ondertekenaar (contactpersoon)", required=True)
    input_cell(ws, "B9", required=True)
    label(ws, 10, "Functie ondertekenaar")
    input_cell(ws, "B10")
    label(ws, 11, "Datum", required=True)
    input_cell(ws, "B11", required=True, comment="Formaat: DD-MM-JJJJ")
    label(ws, 12, "Handtekening", required=True)
    ws.row_dimensions[12].height = 60
    input_cell(ws, "B12", required=True)

    subheader(ws, 14, "Akkoordverklaringen")
    label(ws, 15, "Akkoord met de Cobra-setup zoals beschreven in dit document", required=True)
    dropdown(ws, "B15", "val_JaNee", required=True)
    label(ws, 16, "Akkoord met de schaduwverwerking(en)", required=True)
    dropdown(ws, "B16", "val_JaNee", required=True)

    opmerkingen_block(ws, 18)


# --- Progress formula wiring ---
def wire_progress(wb):
    ws = wb["Start"]
    if not REQUIRED_REFS:
        return
    refs = ",".join(f"'{s}'!{coord_abs(c)}" for s, c in REQUIRED_REFS)
    total = len(REQUIRED_REFS)
    # Filled count / total as text
    ws["B21"].value = f'=COUNTA({refs})&" / {total}"'
    ws["B22"].value = f'=IFERROR(TEXT(COUNTA({refs})/{total},"0%"),"0%")'


def coord_abs(coord):
    """Convert 'B4' -> '$B$4'."""
    import re
    m = re.match(r"([A-Z]+)(\d+)", coord)
    return f"${m.group(1)}${m.group(2)}"


# --- Main ---
def build():
    wb = Workbook()
    wb.remove(wb.active)
    for name in SHEETS:
        ws = wb.create_sheet(name)
        setup_sheet(ws, SHEET_TITLES[name])
    build_validations(wb)
    build_start(wb)
    build_bedrijf(wb)
    build_payroll(wb)
    build_loonheffing(wb)
    build_reserveringen(wb)
    build_verlof_gl(wb)
    build_akkoord(wb)
    wire_progress(wb)
    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    wb.save(OUTPUT)
    return OUTPUT


if __name__ == "__main__":
    path = build()
    size = os.path.getsize(path)
    print(f"SAVED: {path} ({size} bytes, {size/1024:.1f} KB)")
    print(f"Required fields registered: {len(REQUIRED_REFS)}")
