"""QA checks for Payroll_Discovery_Document_MKB.xlsx."""
import os, re, sys
import openpyxl

PATH = "/home/user/nicksSDWorx/outputs/Payroll_Discovery_Document_MKB.xlsx"

def check(label, cond, detail=""):
    mark = "OK" if cond else "FAIL"
    print(f"[{mark}] {label}" + (f"  -- {detail}" if detail else ""))
    return cond

def main():
    ok_all = True

    # 1. File opens (round-trip)
    try:
        wb = openpyxl.load_workbook(PATH)
        ok = True
    except Exception as e:
        ok = False
        print(f"Open failed: {e}")
    ok_all &= check("1. Bestand opent zonder errors", ok)

    # 2. No formula errors stored
    err_count = 0
    for sn in wb.sheetnames:
        for row in wb[sn].iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and re.match(r"^#(REF|VALUE|DIV/0|NAME|N/A|NUM|NULL)!$", v):
                    err_count += 1
    ok_all &= check("2. Geen opgeslagen formule-errors (#REF!, #VALUE! etc.)",
                    err_count == 0, f"{err_count} errors")

    # 3. All DVs point to existing named ranges
    defined = set(wb.defined_names)
    bad_refs = []
    dv_count = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        for dv in ws.data_validations.dataValidation:
            dv_count += 1
            f = dv.formula1 or ""
            if f.startswith("="):
                ref_name = f[1:].strip()
                if ref_name not in defined and not ref_name.startswith("'"):
                    bad_refs.append((sn, ref_name))
    ok_all &= check(
        f"3. Alle {dv_count} dropdowns verwijzen naar bestaande named ranges",
        len(bad_refs) == 0, str(bad_refs) if bad_refs else ""
    )

    # 4. Sheet protection: protected + unlocked input cells sampled
    protection_ok = True
    for sn in wb.sheetnames:
        if sn == "_Validations":
            continue
        if not wb[sn].protection.sheet:
            protection_ok = False
            print(f"   !! {sn} protection.sheet = False")
    # Sample a known input cell and a known label cell
    bedrijf = wb["1. Bedrijf & Contact"]
    input_unlocked = bedrijf["B4"].protection.locked is False
    label_locked = bedrijf["A4"].protection.locked is not False
    ok_all &= check("4a. Alle zichtbare sheets beschermd", protection_ok)
    ok_all &= check("4b. Input-cel '1. Bedrijf & Contact'!B4 unlocked", input_unlocked)
    ok_all &= check("4c. Label-cel '1. Bedrijf & Contact'!A4 locked", label_locked)

    # 5. Total input fields <= 60
    input_count = 0
    for sn in wb.sheetnames:
        if sn == "_Validations":
            continue
        for row in wb[sn].iter_rows():
            for c in row:
                if c.protection.locked is False and c.fill and c.fill.fgColor:
                    rgb = c.fill.fgColor.rgb if c.fill.fgColor else None
                    # Count only yellow (input) cells, not grey defaults not merged areas
                    if rgb in ("00FFF4D1", "FFFFF4D1", "00FFF2CC", "FFFFF2CC"):
                        input_count += 1
    total_unlocked = 0
    for sn in wb.sheetnames:
        if sn == "_Validations":
            continue
        for row in wb[sn].iter_rows():
            for c in row:
                if c.protection.locked is False:
                    total_unlocked += 1
    ok_all &= check(
        f"5. Totaal invulvelden <= 60  (geel-input={input_count}, alle unlocked incl. defaults/checkboxes={total_unlocked})",
        input_count <= 60
    )

    # 6. Every visible sheet: landscape + fitToWidth 1
    page_ok = True
    for sn in wb.sheetnames:
        if sn == "_Validations":
            continue
        ws = wb[sn]
        if ws.page_setup.orientation != "landscape":
            page_ok = False; print(f"   !! {sn} orientation {ws.page_setup.orientation}")
        if ws.page_setup.fitToWidth != 1:
            page_ok = False; print(f"   !! {sn} fitToWidth {ws.page_setup.fitToWidth}")
    ok_all &= check("6. Alle sheets landscape + fitToWidth=1", page_ok)

    # 7. File size <500 KB
    size = os.path.getsize(PATH)
    ok_all &= check(f"7. Bestandsgrootte < 500 KB  ({size/1024:.1f} KB)", size < 500 * 1024)

    # 8. _Validations hidden (not veryHidden)
    state = wb["_Validations"].sheet_state
    ok_all &= check(f"8. _Validations verborgen met state='hidden'  ({state})", state == "hidden")

    # Extra: dropdown resolution sanity check
    for n in sorted(defined):
        dn = wb.defined_names[n]
        if "_Validations" not in dn.attr_text:
            print(f"   ?? named range {n} unexpected attr_text: {dn.attr_text}")

    print("\n" + ("ALL CHECKS PASSED" if ok_all else "SOME CHECKS FAILED"))
    return 0 if ok_all else 1

if __name__ == "__main__":
    sys.exit(main())
