"""
Stamkaart PDF naar Excel Converter

Extracts employee contract and salary history from Dutch HR PDF
("stamkaarten") and exports them to a structured Excel file.

Usage: Double-click the .exe or run `python app.py`
"""

import os
import re
import sys
import logging
import traceback
import threading
from datetime import datetime

import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext


# ---------------------------------------------------------------------------
# Logging setup
# ---------------------------------------------------------------------------

LOG_FILE = os.path.join(os.path.dirname(os.path.abspath(sys.argv[0])), "stamkaart_debug.log")

logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
    ],
)
logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# PDF Parsing
# ---------------------------------------------------------------------------

def extract_text_from_pdf(pdf_path: str) -> list[str]:
    """Extract text from each page of the PDF using pdfplumber."""
    pages_text = []
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            logger.debug(f"=== PAGE {i + 1} RAW TEXT ===\n{text}\n{'=' * 60}")
            pages_text.append(text)
    return pages_text


def parse_date(text: str):
    """Try to parse a Dutch date string into a datetime object."""
    text = text.strip()
    if not text or text == "-":
        return None

    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d %b %Y", "%d-%m-%y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def parse_salary(text: str):
    """Parse a salary string like '2.389,44' or '2389.44' into a float."""
    text = text.strip().replace(" ", "")
    if not text or text == "-":
        return None

    # Dutch format: 2.389,44  ->  remove dots, replace comma with dot
    if "," in text:
        text = text.replace(".", "").replace(",", ".")

    try:
        return round(float(text), 2)
    except ValueError:
        return None


def split_into_employee_blocks(full_text: str) -> list[tuple[str, str]]:
    """
    Split the full PDF text into employee blocks.

    Each employee block typically starts with a line containing the employee
    name followed by stamkaart-related text. We look for patterns like:
      - "Stamkaart" as a section header
      - A name pattern like "Achternaam, V." at the start of a block

    Returns a list of (employee_name, block_text) tuples.
    """
    # Strategy: look for "Stamkaart" headers or lines that contain
    # "Personeelsnummer" which often marks the start of a new employee
    lines = full_text.split("\n")

    # Try to find employee block boundaries using "Stamkaart" keyword
    block_starts = []
    for i, line in enumerate(lines):
        if re.search(r"stamkaart", line, re.IGNORECASE):
            block_starts.append(i)

    if not block_starts:
        # Fallback: treat the entire text as one block, try to find names
        logger.warning("No 'Stamkaart' headers found. Trying alternative splitting.")
        # Look for "Personeelsnummer" as block delimiter
        for i, line in enumerate(lines):
            if re.search(r"personeelsnummer", line, re.IGNORECASE):
                block_starts.append(i)

    if not block_starts:
        # Last resort: treat entire text as one block
        logger.warning("No block delimiters found. Treating entire text as one block.")
        name = extract_employee_name(full_text)
        return [(name or "Onbekend", full_text)]

    blocks = []
    for idx, start in enumerate(block_starts):
        end = block_starts[idx + 1] if idx + 1 < len(block_starts) else len(lines)
        block_text = "\n".join(lines[start:end])
        name = extract_employee_name(block_text)
        if name:
            blocks.append((name, block_text))
        else:
            # Try to get name from the lines just before the "Stamkaart" header
            lookback = max(0, start - 5)
            context = "\n".join(lines[lookback:end])
            name = extract_employee_name(context)
            blocks.append((name or "Onbekend", block_text))

    return blocks


def extract_employee_name(text: str) -> str | None:
    """
    Extract the employee name from a block of text.

    Looks for patterns like:
      - "Naam: Jansen, B."
      - "Achternaam, Voornaam" on its own line
      - "Naam medewerker Jansen, B."
    """
    # Pattern 1: "Naam" or "Naam medewerker" followed by the name
    m = re.search(
        r"(?:naam\s*(?:medewerker)?)\s*[:\-]?\s*([A-Za-zÀ-ÿ\-\s]+,\s*[A-Za-zÀ-ÿ.\-\s]+)",
        text,
        re.IGNORECASE,
    )
    if m:
        return m.group(1).strip()

    # Pattern 2: Look for "Achternaam, X." pattern near the top
    for line in text.split("\n")[:15]:
        line = line.strip()
        m = re.match(r"^([A-Za-zÀ-ÿ\-\s]+,\s*[A-Za-zÀ-ÿ.\-\s]+?)(?:\s{2,}|\t|$)", line)
        if m:
            candidate = m.group(1).strip()
            # Filter out things that are clearly not names
            if len(candidate) > 3 and "," in candidate:
                skip_words = [
                    "contract", "salaris", "mutatie", "datum", "begin",
                    "eind", "dienstverband", "dienstbetrekking", "stamkaart",
                    "personeelsnummer", "functie", "afdeling",
                ]
                if not any(w in candidate.lower() for w in skip_words):
                    return candidate

    return None


def find_section(text: str, section_keyword: str) -> str | None:
    """
    Find a section in the text by keyword (e.g. 'contract mutatie')
    and return the text from that section header to the next section or end.
    """
    lines = text.split("\n")
    start_idx = None

    # Normalize the keyword for flexible matching
    keyword_parts = section_keyword.lower().split()

    for i, line in enumerate(lines):
        line_lower = line.lower()
        # Check if all keyword parts appear in the line
        if all(part in line_lower for part in keyword_parts):
            start_idx = i
            break

    if start_idx is None:
        return None

    # Find the end of this section: next section header or end of text
    # Common section headers in stamkaarten
    section_markers = [
        "mutatie", "historie", "overzicht", "gegevens",
        "stamkaart", "personeelsnummer",
    ]

    end_idx = len(lines)
    for i in range(start_idx + 1, len(lines)):
        line_lower = lines[i].lower().strip()
        # Skip empty lines
        if not line_lower:
            continue
        # Check if this looks like a new section header
        if any(marker in line_lower for marker in section_markers):
            # Make sure it's not just a data row mentioning these words
            # Section headers are typically short and don't have many numbers
            digit_count = sum(1 for c in line_lower if c.isdigit())
            if digit_count < 4:  # Less than 4 digits suggests a header, not data
                end_idx = i
                break

    section_text = "\n".join(lines[start_idx:end_idx])
    logger.debug(f"Found section '{section_keyword}':\n{section_text}")
    return section_text


def parse_contract_rows(section_text: str) -> list[dict]:
    """
    Parse contract mutation rows from the section text.

    Expected columns: Begin contract, Eind contract, Dienstverband, Dienstbetrekking
    """
    rows = []
    lines = section_text.split("\n")

    # Skip the header line(s) - look for lines containing actual date data
    for line in lines:
        line = line.strip()
        if not line:
            continue

        # A contract row should contain at least one date pattern
        date_pattern = r"\d{2}[-/]\d{2}[-/]\d{4}"
        dates = re.findall(date_pattern, line)
        if not dates:
            continue

        # Try to parse the row
        # Expected format: begin_date [end_date] dienstverband dienstbetrekking
        # The end date may be missing (for current contracts)

        # Split by multiple spaces or tabs to get columns
        parts = re.split(r"\s{2,}|\t", line)
        parts = [p.strip() for p in parts if p.strip()]

        if len(parts) < 2:
            # Try splitting differently
            parts = line.split()
            # Reconstruct parts by grouping dates and text
            parts = _smart_split_contract_line(line)

        row = _parse_contract_parts(parts)
        if row:
            rows.append(row)

    return rows


def _smart_split_contract_line(line: str) -> list[str]:
    """
    Intelligently split a contract line into its component parts,
    handling cases where columns may run together.
    """
    parts = []
    date_pattern = r"\d{2}[-/]\d{2}[-/]\d{4}"

    # Find all dates first
    date_matches = list(re.finditer(date_pattern, line))

    if not date_matches:
        return []

    # Extract dates and the text between/after them
    last_end = 0
    for match in date_matches:
        # Text before this date
        before = line[last_end:match.start()].strip()
        if before:
            parts.append(before)
        parts.append(match.group())
        last_end = match.end()

    # Text after the last date
    after = line[last_end:].strip()
    if after:
        # This might contain "Bepaalde tijd Vaste medewerker" etc.
        # Try to split on known keywords
        remaining_parts = _split_contract_text(after)
        parts.extend(remaining_parts)

    return parts


def _split_contract_text(text: str) -> list[str]:
    """Split contract-related text like 'Onbepaalde tijd Vaste medewerker'."""
    text = text.strip()
    if not text:
        return []

    # Known dienstverband values
    dienstverband_patterns = [
        r"(?:On)?bepaalde\s+tijd",
        r"Oproepkracht",
        r"Nul[\s-]?uren",
        r"Min[\s-]?max",
    ]

    for pattern in dienstverband_patterns:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            before = text[:m.start()].strip()
            dienstverband = m.group().strip()
            after = text[m.end():].strip()
            result = []
            if before:
                result.append(before)
            result.append(dienstverband)
            if after:
                result.append(after)
            return result

    return [text]


def _parse_contract_parts(parts: list[str]) -> dict | None:
    """Parse a list of parts into a contract row dict."""
    if len(parts) < 2:
        return None

    row = {
        "begin_contract": None,
        "eind_contract": None,
        "dienstverband": None,
        "dienstbetrekking": None,
    }

    dates_found = []
    texts_found = []

    for part in parts:
        d = parse_date(part)
        if d:
            dates_found.append(d)
        else:
            texts_found.append(part)

    if not dates_found:
        return None

    row["begin_contract"] = dates_found[0] if len(dates_found) >= 1 else None
    row["eind_contract"] = dates_found[1] if len(dates_found) >= 2 else None

    if len(texts_found) >= 1:
        row["dienstverband"] = texts_found[0]
    if len(texts_found) >= 2:
        row["dienstbetrekking"] = texts_found[1]

    return row


def parse_salary_rows(section_text: str) -> list[dict]:
    """
    Parse salary mutation rows from the section text.

    Expected columns: Begin salaris, Eind salaris, Salaris
    """
    rows = []
    lines = section_text.split("\n")

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # A salary row should contain at least one date and a number
        date_pattern = r"\d{2}[-/]\d{2}[-/]\d{4}"
        dates = re.findall(date_pattern, line)
        if not dates:
            continue

        # Split the line
        parts = re.split(r"\s{2,}|\t", line)
        parts = [p.strip() for p in parts if p.strip()]

        if len(parts) < 2:
            parts = _smart_split_salary_line(line)

        row = _parse_salary_parts(parts)
        if row:
            rows.append(row)

    return rows


def _smart_split_salary_line(line: str) -> list[str]:
    """Intelligently split a salary line into its component parts."""
    parts = []
    date_pattern = r"\d{2}[-/]\d{2}[-/]\d{4}"

    date_matches = list(re.finditer(date_pattern, line))
    if not date_matches:
        return []

    last_end = 0
    for match in date_matches:
        before = line[last_end:match.start()].strip()
        if before:
            parts.append(before)
        parts.append(match.group())
        last_end = match.end()

    after = line[last_end:].strip()
    if after:
        # Could contain salary amounts, possibly separated by spaces
        # Salary looks like: 2.389,44 or 2389,44 or 2389.44
        salary_pattern = r"[\d.,]+"
        salary_matches = re.findall(salary_pattern, after)
        for s in salary_matches:
            if len(s) > 2:  # filter out stray digits
                parts.append(s)
        # If no salary matches found, just add the text
        if not salary_matches:
            parts.append(after)

    return parts


def _parse_salary_parts(parts: list[str]) -> dict | None:
    """Parse a list of parts into a salary row dict."""
    if len(parts) < 2:
        return None

    row = {
        "begin_salaris": None,
        "eind_salaris": None,
        "salaris": None,
    }

    dates_found = []
    numbers_found = []

    for part in parts:
        d = parse_date(part)
        if d:
            dates_found.append(d)
        else:
            s = parse_salary(part)
            if s is not None:
                numbers_found.append(s)

    if not dates_found:
        return None

    row["begin_salaris"] = dates_found[0] if len(dates_found) >= 1 else None
    row["eind_salaris"] = dates_found[1] if len(dates_found) >= 2 else None

    # The salary amount is typically the last number
    if numbers_found:
        row["salaris"] = numbers_found[-1]

    return row


def process_pdf(pdf_path: str, progress_callback=None) -> list[dict]:
    """
    Main processing function: extract all employee data from the PDF.

    Returns a list of row dicts ready for Excel export.
    """
    def report(msg):
        logger.info(msg)
        if progress_callback:
            progress_callback(msg)

    report(f"PDF openen: {pdf_path}")

    # Step 1: Extract text from all pages
    pages = extract_text_from_pdf(pdf_path)
    report(f"{len(pages)} pagina('s) gevonden.")

    # Combine all pages into one text block
    full_text = "\n".join(pages)

    # Step 2: Split into employee blocks
    employee_blocks = split_into_employee_blocks(full_text)
    report(f"{len(employee_blocks)} medewerker(s) gevonden.")

    all_rows = []
    warnings = []

    for emp_name, block_text in employee_blocks:
        report(f"Verwerken: {emp_name}")
        logger.debug(f"Employee block for '{emp_name}':\n{block_text[:500]}...")

        # Extract contract mutations
        contract_section = find_section(block_text, "contract mutatie")
        if contract_section is None:
            # Also try alternative keywords
            contract_section = find_section(block_text, "contract historie")
        if contract_section is None:
            contract_section = find_section(block_text, "contractgegevens")

        if contract_section:
            contract_rows = parse_contract_rows(contract_section)
            report(f"  {len(contract_rows)} contractregel(s) gevonden.")
            for crow in contract_rows:
                all_rows.append({
                    "Naam": emp_name,
                    "Begin contract": crow["begin_contract"],
                    "Eind contract": crow["eind_contract"],
                    "Dienstverband": crow["dienstverband"],
                    "Dienstbetrekking": crow["dienstbetrekking"],
                    "Begin salaris": None,
                    "Eind salaris": None,
                    "Salaris": None,
                })
        else:
            msg = f"  WAARSCHUWING: Geen 'Contract mutaties' gevonden voor {emp_name}."
            report(msg)
            warnings.append(msg)

        # Extract salary mutations
        salary_section = find_section(block_text, "salaris mutatie")
        if salary_section is None:
            salary_section = find_section(block_text, "salaris historie")
        if salary_section is None:
            salary_section = find_section(block_text, "salarisgegevens")

        if salary_section:
            salary_rows = parse_salary_rows(salary_section)
            report(f"  {len(salary_rows)} salarisregel(s) gevonden.")
            for srow in salary_rows:
                all_rows.append({
                    "Naam": emp_name,
                    "Begin contract": None,
                    "Eind contract": None,
                    "Dienstverband": None,
                    "Dienstbetrekking": None,
                    "Begin salaris": srow["begin_salaris"],
                    "Eind salaris": srow["eind_salaris"],
                    "Salaris": srow["salaris"],
                })
        else:
            msg = f"  WAARSCHUWING: Geen 'Salaris mutaties' gevonden voor {emp_name}."
            report(msg)
            warnings.append(msg)

    report(f"\nTotaal: {len(all_rows)} rijen geëxtraheerd.")
    if warnings:
        report(f"\n{len(warnings)} waarschuwing(en):")
        for w in warnings:
            report(f"  {w}")

    return all_rows


# ---------------------------------------------------------------------------
# Excel Export
# ---------------------------------------------------------------------------

COLUMNS = [
    "Naam",
    "Begin contract",
    "Eind contract",
    "Dienstverband",
    "Dienstbetrekking",
    "Begin salaris",
    "Eind salaris",
    "Salaris",
]

DATE_COLUMNS = {"Begin contract", "Eind contract", "Begin salaris", "Eind salaris"}


def write_excel(rows: list[dict], output_path: str):
    """Write the extracted data to an Excel file matching the expected format."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Write data rows
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            value = row_data.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            if col_name in DATE_COLUMNS and value is not None:
                cell.number_format = "YYYY-MM-DD"
            elif col_name == "Salaris" and value is not None:
                cell.number_format = "#,##0.00"

    # Auto-fit column widths (approximate)
    col_widths = {
        "Naam": 20,
        "Begin contract": 15,
        "Eind contract": 15,
        "Dienstverband": 18,
        "Dienstbetrekking": 20,
        "Begin salaris": 15,
        "Eind salaris": 15,
        "Salaris": 12,
    }
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_widths.get(col_name, 15)

    wb.save(output_path)
    logger.info(f"Excel bestand opgeslagen: {output_path}")


# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------

class StamkaartApp:
    """Dutch-language tkinter GUI for the PDF-to-Excel converter."""

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Stamkaart PDF naar Excel")
        self.root.geometry("700x520")
        self.root.resizable(True, True)

        # Try to set a modern look
        try:
            self.root.tk.call("tk", "scaling", 1.2)
        except tk.TclError:
            pass

        self.pdf_path = tk.StringVar()
        self.xlsx_path = tk.StringVar()

        self._build_ui()

    def _build_ui(self):
        # Main frame with padding
        main = tk.Frame(self.root, padx=15, pady=15)
        main.pack(fill=tk.BOTH, expand=True)

        # Title
        title = tk.Label(
            main,
            text="Stamkaart PDF naar Excel Converter",
            font=("Segoe UI", 14, "bold"),
        )
        title.pack(pady=(0, 15))

        # --- PDF input ---
        pdf_frame = tk.LabelFrame(main, text="Invoer", padx=10, pady=8)
        pdf_frame.pack(fill=tk.X, pady=(0, 8))

        tk.Label(pdf_frame, text="PDF bestand:").grid(row=0, column=0, sticky="w")
        tk.Entry(pdf_frame, textvariable=self.pdf_path, width=55).grid(
            row=0, column=1, padx=5
        )
        tk.Button(pdf_frame, text="Bladeren...", command=self._browse_pdf).grid(
            row=0, column=2
        )

        # --- Excel output ---
        xlsx_frame = tk.LabelFrame(main, text="Uitvoer", padx=10, pady=8)
        xlsx_frame.pack(fill=tk.X, pady=(0, 8))

        tk.Label(xlsx_frame, text="Excel bestand:").grid(row=0, column=0, sticky="w")
        tk.Entry(xlsx_frame, textvariable=self.xlsx_path, width=55).grid(
            row=0, column=1, padx=5
        )
        tk.Button(xlsx_frame, text="Bladeren...", command=self._browse_xlsx).grid(
            row=0, column=2
        )

        # --- Execute button ---
        btn_frame = tk.Frame(main)
        btn_frame.pack(fill=tk.X, pady=8)

        self.run_btn = tk.Button(
            btn_frame,
            text="Uitvoeren",
            command=self._run,
            font=("Segoe UI", 11, "bold"),
            bg="#4CAF50",
            fg="white",
            padx=20,
            pady=5,
        )
        self.run_btn.pack()

        # --- Status area ---
        status_frame = tk.LabelFrame(main, text="Status", padx=10, pady=8)
        status_frame.pack(fill=tk.BOTH, expand=True, pady=(8, 0))

        self.status_text = scrolledtext.ScrolledText(
            status_frame, height=12, state=tk.DISABLED, wrap=tk.WORD,
            font=("Consolas", 9),
        )
        self.status_text.pack(fill=tk.BOTH, expand=True)

    def _browse_pdf(self):
        path = filedialog.askopenfilename(
            title="Selecteer PDF bestand",
            filetypes=[("PDF bestanden", "*.pdf"), ("Alle bestanden", "*.*")],
        )
        if path:
            self.pdf_path.set(path)
            # Auto-suggest output path
            if not self.xlsx_path.get():
                base = os.path.splitext(path)[0]
                self.xlsx_path.set(base + "_export.xlsx")

    def _browse_xlsx(self):
        path = filedialog.asksaveasfilename(
            title="Opslaan als Excel bestand",
            defaultextension=".xlsx",
            filetypes=[("Excel bestanden", "*.xlsx"), ("Alle bestanden", "*.*")],
        )
        if path:
            self.xlsx_path.set(path)

    def _log(self, message: str):
        """Append a message to the status text area (thread-safe)."""
        def _update():
            self.status_text.config(state=tk.NORMAL)
            self.status_text.insert(tk.END, message + "\n")
            self.status_text.see(tk.END)
            self.status_text.config(state=tk.DISABLED)
        self.root.after(0, _update)

    def _run(self):
        """Start the extraction in a background thread."""
        pdf = self.pdf_path.get().strip()
        xlsx = self.xlsx_path.get().strip()

        if not pdf:
            messagebox.showwarning("Waarschuwing", "Selecteer eerst een PDF bestand.")
            return
        if not os.path.isfile(pdf):
            messagebox.showerror("Fout", f"PDF bestand niet gevonden:\n{pdf}")
            return
        if not xlsx:
            messagebox.showwarning("Waarschuwing", "Kies een locatie voor het Excel bestand.")
            return

        # Clear status
        self.status_text.config(state=tk.NORMAL)
        self.status_text.delete("1.0", tk.END)
        self.status_text.config(state=tk.DISABLED)

        self.run_btn.config(state=tk.DISABLED, text="Bezig...")

        thread = threading.Thread(target=self._extract, args=(pdf, xlsx), daemon=True)
        thread.start()

    def _extract(self, pdf_path: str, xlsx_path: str):
        """Run the extraction (called in background thread)."""
        try:
            rows = process_pdf(pdf_path, progress_callback=self._log)

            if not rows:
                self._log("\nGeen gegevens gevonden in de PDF.")
                self._log("Controleer het debug logbestand voor meer informatie:")
                self._log(f"  {LOG_FILE}")
                self.root.after(0, lambda: messagebox.showwarning(
                    "Waarschuwing",
                    "Geen gegevens gevonden. Controleer het logbestand voor details."
                ))
                return

            write_excel(rows, xlsx_path)
            self._log(f"\nExcel bestand opgeslagen: {xlsx_path}")
            self._log("Klaar!")
            self.root.after(0, lambda: messagebox.showinfo(
                "Gereed",
                f"Export voltooid!\n\n{len(rows)} rijen geschreven naar:\n{xlsx_path}"
            ))

        except Exception as e:
            error_msg = traceback.format_exc()
            logger.error(f"Error during extraction:\n{error_msg}")
            self._log(f"\nFOUT: {e}")
            self._log(f"\nDetails in logbestand: {LOG_FILE}")
            self.root.after(0, lambda: messagebox.showerror(
                "Fout",
                f"Er is een fout opgetreden:\n{e}\n\nZie logbestand voor details."
            ))

        finally:
            self.root.after(0, lambda: self.run_btn.config(
                state=tk.NORMAL, text="Uitvoeren"
            ))


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    root = tk.Tk()
    StamkaartApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
