"""Wekelijks voortgangsrapport-generator.

Leest een implementatie-tracking Excel READ-ONLY in, bouwt een snapshot,
vergelijkt met de vorige snapshot, en genereert een Markdown-rapport.

Compliance:
- Geen netwerk-calls. Alleen pandas, openpyxl en stdlib.
- Excel wordt nooit gewijzigd (read_only=True).
- Snapshots als leesbare JSON (geen pickle).
- Logs bevatten geen klantdata; alleen aggregaten en bestandsnamen.
"""

from __future__ import annotations

import json
import os
import sys
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent
CONFIG_PATH = PROJECT_ROOT / "config.json"
SNAPSHOT_DIR = PROJECT_ROOT / "data" / "snapshots"
REPORT_DIR = PROJECT_ROOT / "data" / "reports"


# ---------------------------------------------------------------------------
# Hulpfuncties
# ---------------------------------------------------------------------------

def log(msg: str) -> None:
    """Print zonder klantdata. Alleen aggregaten/bestandsnamen/foutsoorten."""
    print(f"[INFO] {msg}")


def err(msg: str) -> None:
    print(f"[FOUT] {msg}", file=sys.stderr)


def is_empty(value: Any, empty_set: list) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    if isinstance(value, str) and value.strip() in {str(x).strip() for x in empty_set if x is not None}:
        return True
    return False


def to_iso_date(value: Any) -> str | None:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, (datetime, date)):
        d = value.date() if isinstance(value, datetime) else value
        return d.isoformat()
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                continue
        return None
    return None


def parse_iso_date(s: str | None) -> date | None:
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def safe_str(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return str(value).strip()


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

@dataclass
class Config:
    raw: dict

    def get(self, *path: str, default: Any = None) -> Any:
        node: Any = self.raw
        for key in path:
            if not isinstance(node, dict) or key not in node:
                return default
            node = node[key]
        return node


def load_config() -> Config:
    if not CONFIG_PATH.exists():
        raise FileNotFoundError(
            f"Configuratiebestand niet gevonden: {CONFIG_PATH.name}. "
            "Plaats config.json in de projectmap."
        )
    try:
        with CONFIG_PATH.open("r", encoding="utf-8") as f:
            raw = json.load(f)
    except json.JSONDecodeError as e:
        raise ValueError(
            f"config.json bevat een syntaxfout op regel {e.lineno}. "
            "Controleer komma's en aanhalingstekens."
        ) from None
    required = ["excel_path", "general_sheet", "columns", "customer_sheet"]
    for key in required:
        if key not in raw:
            raise ValueError(f"config.json mist verplicht veld: '{key}'.")
    return Config(raw)


# ---------------------------------------------------------------------------
# Excel inlezen
# ---------------------------------------------------------------------------

def read_general(excel_path: Path, cfg: Config) -> pd.DataFrame:
    sheet = cfg.get("general_sheet", default="GENERAL")
    header_row = cfg.get("general_header_row", default=1)
    try:
        df = pd.read_excel(
            excel_path,
            sheet_name=sheet,
            header=header_row - 1,
            engine="openpyxl",
        )
    except ValueError as e:
        raise ValueError(
            f"Kan tabblad '{sheet}' niet lezen. "
            f"Bestaat het tabblad nog in {excel_path.name}?"
        ) from e
    df = df.dropna(how="all")
    return df


def read_customer_sheet(wb, sheet_name: str, cfg: Config) -> dict | None:
    ws = wb[sheet_name]
    cust_cfg = cfg.get("customer_sheet", default={})
    meta_cells = cust_cfg.get("metadata_cells", {})
    scorecard = cust_cfg.get("phase_scorecard", {})
    header_row = cust_cfg.get("header_row", 12)
    task_cols = cust_cfg.get("task_columns", {})
    skip_hide = set(cust_cfg.get("skip_when_hide_in", []))

    metadata = {}
    for key, cell_ref in meta_cells.items():
        try:
            value = ws[cell_ref].value
        except (ValueError, KeyError):
            value = None
        if isinstance(value, datetime):
            metadata[key] = value.date().isoformat()
        elif isinstance(value, date):
            metadata[key] = value.isoformat()
        else:
            metadata[key] = safe_str(value) or None

    phase_scorecard = {}
    for phase, cell_ref in scorecard.items():
        try:
            value = ws[cell_ref].value
        except (ValueError, KeyError):
            value = None
        if isinstance(value, (int, float)) and not pd.isna(value):
            phase_scorecard[phase] = float(value)
        else:
            phase_scorecard[phase] = None

    header_values = {}
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        if cell.value:
            header_values[safe_str(cell.value)] = col_idx

    def col(name: str) -> int | None:
        return header_values.get(name)

    taak_col = col(task_cols.get("taak", "Taak"))
    pct_col = col(task_cols.get("status_pct", "Status %"))
    hide_col = col(task_cols.get("hide", "Hide"))

    task_count = 0
    tasks_done = 0
    tasks_pct_sum = 0.0
    tasks_pct_count = 0

    if taak_col and pct_col:
        for row_idx in range(header_row + 1, ws.max_row + 1):
            taak_value = ws.cell(row=row_idx, column=taak_col).value
            if not taak_value or not safe_str(taak_value):
                continue
            if hide_col is not None:
                hide_value = ws.cell(row=row_idx, column=hide_col).value
                if hide_value in skip_hide:
                    continue
            task_count += 1
            pct_value = ws.cell(row=row_idx, column=pct_col).value
            if isinstance(pct_value, (int, float)) and not pd.isna(pct_value):
                pct_float = float(pct_value)
                if pct_float <= 1.5:
                    pct_float *= 100.0
                tasks_pct_sum += pct_float
                tasks_pct_count += 1
                if pct_float >= 99.5:
                    tasks_done += 1

    tasks_avg_pct = round(tasks_pct_sum / tasks_pct_count, 1) if tasks_pct_count else 0.0

    return {
        "sheet_name": sheet_name,
        "metadata": metadata,
        "scorecard": phase_scorecard,
        "task_count": task_count,
        "tasks_done": tasks_done,
        "tasks_avg_pct": tasks_avg_pct,
    }


def read_excel(cfg: Config) -> tuple[pd.DataFrame, dict[str, dict]]:
    excel_path = (PROJECT_ROOT / cfg.get("excel_path", default="dummy_tracker.xlsx")).resolve()
    if not excel_path.exists():
        raise FileNotFoundError(
            f"Excel-bestand niet gevonden: {excel_path.name}. "
            f"Controleer 'excel_path' in config.json."
        )
    log(f"Lees bestand: {excel_path.name}")
    general_df = read_general(excel_path, cfg)
    log(f"GENERAL ingelezen: {len(general_df)} rijen")

    skip_sheets = set(cfg.get("skip_sheets", default=[]))
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    customer_tabs: dict[str, dict] = {}
    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue
        try:
            data = read_customer_sheet(wb, sheet_name, cfg)
        except Exception as e:
            log(f"Tabblad overgeslagen ({type(e).__name__}): {sheet_name}")
            continue
        if data is None:
            continue
        affiliate = data["metadata"].get("affiliate")
        key = safe_str(affiliate) or sheet_name
        customer_tabs[key] = data
    wb.close()
    log(f"Klant-tabbladen verwerkt: {len(customer_tabs)}")
    return general_df, customer_tabs


# ---------------------------------------------------------------------------
# Snapshot bouwen
# ---------------------------------------------------------------------------

def derive_overall_status(impl: dict, cust_tab: dict | None, cfg: Config) -> str:
    empty_set = cfg.get("status_definition", "treat_as_empty", default=[" ", ""])
    chain = [cfg.get("status_definition", "primary_source", default="general_overall_status")]
    chain += cfg.get("status_definition", "fallback_chain", default=[])

    for source in chain:
        if source == "general_overall_status":
            value = impl.get("overall_status")
            if not is_empty(value, empty_set):
                return safe_str(value)
        elif source == "customer_tab_overall":
            if cust_tab:
                overall = cust_tab.get("scorecard", {}).get("Overall")
                if isinstance(overall, (int, float)):
                    pct = overall * 100 if overall <= 1.5 else overall
                    if pct >= 99.5:
                        return "Afgerond"
                    if pct > 0:
                        return f"In progress ({pct:.0f}%)"
        elif source == "task_avg_pct":
            if cust_tab:
                avg = cust_tab.get("tasks_avg_pct", 0)
                if avg >= 99.5:
                    return "Afgerond"
                if avg > 0:
                    return f"In progress ({avg:.0f}%)"
    return "Onbekend"


def build_snapshot(general_df: pd.DataFrame, customer_tabs: dict, cfg: Config, today: date) -> dict:
    cols = cfg.get("columns", default={})
    col_kn = cols.get("klantnummer")
    implementations: dict[str, dict] = {}

    for _, row in general_df.iterrows():
        kn_raw = row.get(col_kn)
        if kn_raw is None or (isinstance(kn_raw, float) and pd.isna(kn_raw)):
            continue
        kn = safe_str(kn_raw).rstrip(".0") if isinstance(kn_raw, float) else safe_str(kn_raw)
        if not kn:
            continue

        impl: dict[str, Any] = {
            "customer": safe_str(row.get(cols.get("customer"))) or None,
            "owner": safe_str(row.get(cols.get("owner"))) or None,
            "service": safe_str(row.get(cols.get("service"))) or None,
            "hrm": safe_str(row.get(cols.get("hrm"))) or None,
            "go_live": to_iso_date(row.get(cols.get("go_live"))),
            "overall_status": safe_str(row.get(cols.get("overall_status"))) or None,
            "aanleverdatum": safe_str(row.get(cols.get("aanleverdatum"))) or None,
            "eerste_dag_hrm": safe_str(row.get(cols.get("eerste_dag_hrm"))) or None,
            "wns": _to_int(row.get(cols.get("wns"))),
            "laatste_update": safe_str(row.get(cols.get("laatste_update"))) or None,
            "phases": {},
        }
        for phase in cols.get("phase_columns", []) or []:
            impl["phases"][phase] = safe_str(row.get(phase)) or None

        cust_tab = customer_tabs.get(kn)
        impl["customer_tab"] = cust_tab if cust_tab else None
        impl["status"] = derive_overall_status(impl, cust_tab, cfg)
        implementations[kn] = impl

    return {
        "snapshot_date": today.isoformat(),
        "implementation_count": len(implementations),
        "implementations": implementations,
    }


def _to_int(value: Any) -> int | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


# ---------------------------------------------------------------------------
# Snapshot opslaan & laden
# ---------------------------------------------------------------------------

def save_snapshot(snapshot: dict) -> Path:
    SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
    path = SNAPSHOT_DIR / f"{snapshot['snapshot_date']}_snapshot.json"
    with path.open("w", encoding="utf-8") as f:
        json.dump(snapshot, f, indent=2, ensure_ascii=False, default=str)
    log(f"Snapshot opgeslagen: {path.name} ({snapshot['implementation_count']} implementaties)")
    return path


def list_snapshots() -> list[Path]:
    if not SNAPSHOT_DIR.exists():
        return []
    return sorted(SNAPSHOT_DIR.glob("*_snapshot.json"))


def load_previous_snapshot(today_iso: str) -> dict | None:
    candidates = [p for p in list_snapshots() if p.stem.split("_")[0] < today_iso]
    if not candidates:
        return None
    latest = candidates[-1]
    try:
        with latest.open("r", encoding="utf-8") as f:
            return json.load(f)
    except (OSError, json.JSONDecodeError) as e:
        err(f"Vorige snapshot onleesbaar ({type(e).__name__}): {latest.name}")
        return None


# ---------------------------------------------------------------------------
# Diff
# ---------------------------------------------------------------------------

@dataclass
class Diff:
    new: list[str] = field(default_factory=list)
    removed: list[str] = field(default_factory=list)
    status_changed: list[tuple[str, str, str]] = field(default_factory=list)
    unchanged_active: list[str] = field(default_factory=list)
    completed: list[str] = field(default_factory=list)


def compute_diff(current: dict, previous: dict | None, cfg: Config) -> Diff:
    diff = Diff()
    if previous is None:
        return diff

    cur_impls = current["implementations"]
    prev_impls = previous.get("implementations", {})
    done_vals = {str(v).lower() for v in cfg.get("status_definition", "done_values", default=[])}

    for kn, impl in cur_impls.items():
        if kn not in prev_impls:
            diff.new.append(kn)
            continue
        prev = prev_impls[kn]
        cur_status = impl.get("status") or ""
        prev_status = prev.get("status") or ""
        if cur_status != prev_status:
            diff.status_changed.append((kn, prev_status, cur_status))
            if any(d in cur_status.lower() for d in done_vals):
                diff.completed.append(kn)
        else:
            is_done = any(d in cur_status.lower() for d in done_vals)
            is_unknown = cur_status.lower() == "onbekend"
            if not is_done and not is_unknown:
                diff.unchanged_active.append(kn)

    for kn in prev_impls:
        if kn not in cur_impls:
            diff.removed.append(kn)

    return diff


# ---------------------------------------------------------------------------
# Knelpunten
# ---------------------------------------------------------------------------

@dataclass
class Bottlenecks:
    deadline_near: list[str] = field(default_factory=list)
    deadline_overdue: list[str] = field(default_factory=list)
    missing_owner: list[str] = field(default_factory=list)
    missing_deadline: list[str] = field(default_factory=list)
    critical_date_ntb: list[str] = field(default_factory=list)
    stagnation: list[tuple[str, int]] = field(default_factory=list)
    no_task_progress: list[str] = field(default_factory=list)


def detect_bottlenecks(snapshot: dict, cfg: Config, today: date) -> Bottlenecks:
    b = Bottlenecks()
    th = cfg.get("thresholds", default={})
    done_vals = {str(v).lower() for v in cfg.get("status_definition", "done_values", default=[])}
    ntb_vals = {str(v).lower() for v in th.get("ntb_values", [])}
    warning_days = th.get("deadline_warning_days", 7)
    critical_window = th.get("critical_date_window_days", 60)
    no_progress_window = th.get("no_progress_window_days", 90)
    stagnation_weeks = th.get("stagnation_weeks", 3)

    def is_done(status: str) -> bool:
        s = (status or "").lower()
        return any(d in s for d in done_vals)

    for kn, impl in snapshot["implementations"].items():
        status = impl.get("status") or ""
        if is_done(status):
            continue

        go_live = parse_iso_date(impl.get("go_live"))
        if go_live:
            delta = (go_live - today).days
            if delta < 0:
                b.deadline_overdue.append(kn)
            elif delta <= warning_days:
                b.deadline_near.append(kn)
        else:
            if th.get("flag_missing_deadline", True):
                b.missing_deadline.append(kn)

        if th.get("flag_missing_owner", True) and not impl.get("owner"):
            b.missing_owner.append(kn)

        if go_live and 0 <= (go_live - today).days <= critical_window:
            for field_name in ("aanleverdatum", "eerste_dag_hrm"):
                value = (impl.get(field_name) or "").lower()
                if value in ntb_vals:
                    b.critical_date_ntb.append(kn)
                    break

    history = _load_snapshot_history()
    if history:
        for kn, impl in snapshot["implementations"].items():
            weeks = _status_unchanged_weeks(kn, impl.get("status"), history, today)
            if weeks >= stagnation_weeks and not is_done(impl.get("status") or ""):
                b.stagnation.append((kn, weeks))

    if history and len(history) >= 2:
        previous = history[-1] if history[-1].get("snapshot_date") != snapshot.get("snapshot_date") else (history[-2] if len(history) >= 2 else None)
        if previous:
            prev_impls = previous.get("implementations", {})
            for kn, impl in snapshot["implementations"].items():
                if is_done(impl.get("status") or ""):
                    continue
                go_live = parse_iso_date(impl.get("go_live"))
                if not go_live or (go_live - today).days > no_progress_window:
                    continue
                cur_tab = impl.get("customer_tab") or {}
                prev_tab = (prev_impls.get(kn) or {}).get("customer_tab") or {}
                if cur_tab.get("tasks_avg_pct") is not None and \
                   cur_tab.get("tasks_avg_pct") == prev_tab.get("tasks_avg_pct"):
                    b.no_task_progress.append(kn)

    return b


def _load_snapshot_history() -> list[dict]:
    history = []
    for path in list_snapshots():
        try:
            with path.open("r", encoding="utf-8") as f:
                history.append(json.load(f))
        except (OSError, json.JSONDecodeError):
            continue
    return history


def _status_unchanged_weeks(kn: str, current_status: str | None, history: list[dict], today: date) -> int:
    if not current_status:
        return 0
    earliest_same: date | None = None
    for snap in history:
        snap_date = parse_iso_date(snap.get("snapshot_date"))
        if not snap_date:
            continue
        impl = snap.get("implementations", {}).get(kn)
        if not impl:
            continue
        if impl.get("status") == current_status:
            if earliest_same is None or snap_date < earliest_same:
                earliest_same = snap_date
        else:
            earliest_same = None
    if not earliest_same:
        return 0
    return (today - earliest_same).days // 7


# ---------------------------------------------------------------------------
# Rapport
# ---------------------------------------------------------------------------

def render_report(snapshot: dict, diff: Diff, bottlenecks: Bottlenecks,
                  previous: dict | None, cfg: Config, today: date) -> Path:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    path = REPORT_DIR / f"{today.isoformat()}_voortgangsrapport.md"

    impls = snapshot["implementations"]
    status_totals: dict[str, int] = {}
    for impl in impls.values():
        status = impl.get("status") or "Onbekend"
        key = status.split(" (")[0]
        status_totals[key] = status_totals.get(key, 0) + 1

    lines: list[str] = []
    lines.append(f"# Voortgangsrapport — {today.isoformat()}")
    lines.append("")

    # Samenvatting
    lines.append("## Samenvatting")
    lines.append("")
    lines.append(f"- Totaal implementaties: **{len(impls)}**")
    lines.append("- Per status:")
    for key in sorted(status_totals):
        lines.append(f"  - {key}: {status_totals[key]}")
    if previous:
        prev_date = previous.get("snapshot_date", "onbekend")
        lines.append("")
        lines.append(f"Vergelijking met snapshot van **{prev_date}**:")
        lines.append(f"- Nieuw: {len(diff.new)}")
        lines.append(f"- Status-wijzigingen: {len(diff.status_changed)}")
        lines.append(f"- Afgerond: {len(diff.completed)}")
        lines.append(f"- Verdwenen uit lijst: {len(diff.removed)}")
    else:
        lines.append("")
        lines.append("> Geen vorige snapshot beschikbaar — dit is de eerste week. Diff-secties zijn leeg.")
    lines.append("")

    # Status per implementatie
    lines.append("## Status per implementatie")
    lines.append("")
    lines.append("| Klantnr | Customer | Owner | Go-live | Status | # wns |")
    lines.append("|---|---|---|---|---|---|")
    for kn in sorted(impls):
        impl = impls[kn]
        lines.append(
            f"| {kn} | {impl.get('customer') or '—'} | {impl.get('owner') or '—'} | "
            f"{impl.get('go_live') or '—'} | {impl.get('status') or '—'} | {impl.get('wns') or '—'} |"
        )
    lines.append("")

    # Wijzigingen
    lines.append("## Wijzigingen sinds vorige week")
    lines.append("")
    _section(lines, "### Nieuwe implementaties", diff.new, impls)
    if diff.status_changed:
        lines.append("### Status-wijzigingen")
        lines.append("")
        for kn, old, new in diff.status_changed:
            cust = impls.get(kn, {}).get("customer") or "—"
            lines.append(f"- **{kn}** ({cust}): `{old or 'leeg'}` → `{new or 'leeg'}`")
        lines.append("")
    else:
        lines.append("### Status-wijzigingen")
        lines.append("")
        lines.append("_Geen status-wijzigingen._")
        lines.append("")
    _section(lines, "### Afgerond deze week", diff.completed, impls)
    _section(lines, "### Mogelijk stagnerend (zelfde status sinds vorige week)", diff.unchanged_active, impls, limit=20)
    _section(lines, "### Verwijderd / niet meer in lijst", diff.removed, {})

    # Knelpunten
    lines.append("## Knelpunten en signalen")
    lines.append("")
    _section(lines, "### Deadlines binnen 7 dagen", bottlenecks.deadline_near, impls)
    _section(lines, "### Verstreken deadlines", bottlenecks.deadline_overdue, impls)
    _section(lines, "### Ontbrekende eigenaar", bottlenecks.missing_owner, impls)
    _section(lines, "### Ontbrekende deadline", bottlenecks.missing_deadline, impls, limit=15)
    _section(lines, "### Kritieke datum nog op 'Ntb'", bottlenecks.critical_date_ntb, impls)
    if bottlenecks.stagnation:
        lines.append("### Stagnatie (≥ 3 weken zelfde status)")
        lines.append("")
        for kn, weeks in bottlenecks.stagnation:
            cust = impls.get(kn, {}).get("customer") or "—"
            lines.append(f"- **{kn}** ({cust}): {weeks} weken zelfde status")
        lines.append("")
    _section(lines, "### Geen taak-voortgang (binnen 90 dagen Go-live)", bottlenecks.no_task_progress, impls)

    # Per-klant blokken
    if cfg.get("report", "include_per_customer_blocks", default=True):
        relevant_only = cfg.get("report", "per_customer_only_with_changes_or_bottlenecks", default=True)
        flagged = _collect_flagged_keys(diff, bottlenecks)
        keys_for_blocks = sorted(flagged) if relevant_only else sorted(impls)
        if keys_for_blocks:
            lines.append("## Per implementatie — detail")
            lines.append("")
            for kn in keys_for_blocks:
                impl = impls.get(kn)
                if not impl:
                    continue
                _render_customer_block(lines, kn, impl)

    content = "\n".join(lines) + "\n"
    path.write_text(content, encoding="utf-8")
    log(f"Rapport geschreven: {path.name}")
    return path


def _section(lines: list[str], heading: str, keys: list[str], impls: dict, limit: int = 30) -> None:
    lines.append(heading)
    lines.append("")
    if not keys:
        lines.append("_Geen._")
        lines.append("")
        return
    shown = keys[:limit]
    for kn in shown:
        impl = impls.get(kn, {})
        cust = impl.get("customer") or "—"
        lines.append(f"- **{kn}** ({cust})")
    if len(keys) > limit:
        lines.append(f"- _… en {len(keys) - limit} meer (zie snapshot voor volledige lijst)_")
    lines.append("")


def _collect_flagged_keys(diff: Diff, b: Bottlenecks) -> set[str]:
    flagged: set[str] = set()
    flagged.update(diff.new, diff.completed)
    flagged.update(kn for kn, _, _ in diff.status_changed)
    flagged.update(b.deadline_near, b.deadline_overdue, b.missing_owner,
                   b.missing_deadline, b.critical_date_ntb, b.no_task_progress)
    flagged.update(kn for kn, _ in b.stagnation)
    return flagged


def _render_customer_block(lines: list[str], kn: str, impl: dict) -> None:
    lines.append(f"### {kn} — {impl.get('customer') or '—'}")
    lines.append("")
    lines.append(f"- Owner: {impl.get('owner') or '—'}")
    lines.append(f"- Service: {impl.get('service') or '—'}")
    lines.append(f"- Go-live: {impl.get('go_live') or '—'}")
    lines.append(f"- Status: {impl.get('status') or '—'}")
    lines.append(f"- # wns: {impl.get('wns') or '—'}")
    laatste = impl.get("laatste_update")
    if laatste and laatste not in ("0", "-"):
        lines.append(f"- Laatste update: {laatste}")
    tab = impl.get("customer_tab")
    if tab:
        sc = tab.get("scorecard") or {}
        scorecard_parts = [f"{k}: {int(v) if isinstance(v, (int, float)) and not pd.isna(v) else '—'}"
                           for k, v in sc.items()]
        if scorecard_parts:
            lines.append(f"- Scorecard: {', '.join(scorecard_parts)}")
        lines.append(
            f"- Taken: {tab.get('tasks_done', 0)} afgerond / {tab.get('task_count', 0)} totaal "
            f"(gem. {tab.get('tasks_avg_pct', 0):.0f}%)"
        )
    lines.append("")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def open_report(path: Path, cfg: Config) -> None:
    if not cfg.get("report", "open_after_generate", default=True):
        return
    try:
        if hasattr(os, "startfile"):
            os.startfile(str(path))
        else:
            log(f"Rapport beschikbaar op: {path}")
    except OSError as e:
        log(f"Rapport openen mislukt ({type(e).__name__}). Open handmatig: {path}")


def main() -> int:
    try:
        cfg = load_config()
        today = date.today()
        general_df, customer_tabs = read_excel(cfg)
        snapshot = build_snapshot(general_df, customer_tabs, cfg, today)
        previous = load_previous_snapshot(today.isoformat())
        diff = compute_diff(snapshot, previous, cfg)
        save_snapshot(snapshot)
        bottlenecks = detect_bottlenecks(snapshot, cfg, today)
        report_path = render_report(snapshot, diff, bottlenecks, previous, cfg, today)
        open_report(report_path, cfg)
        log("Klaar.")
        return 0
    except FileNotFoundError as e:
        err(str(e))
        return 1
    except ValueError as e:
        err(str(e))
        return 1
    except Exception as e:
        err(f"Onverwachte fout ({type(e).__name__}). Details: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
